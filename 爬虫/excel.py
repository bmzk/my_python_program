import openpyxl
from tkinter import filedialog
# 常量 #################################


# 全局变量 ##############################
省份位置 = 0
表格名称列表 = []
行号=0
##########################################
class 工作簿类():
    def __init__(self,表格名称列表):
        self.工作簿 = openpyxl.Workbook()
        # 创建新工作表
        for i in 表格名称列表:
            self.工作簿.create_sheet(i) 
        self.工作簿.close()
    def 写入一行数据(self,表格名称:str,行数据:list):
        self.工作簿[表格名称].append(行数据)
        #print(self.省名,表格名称,'写入数据 ',行数据)


def 保存工作簿文件(工作簿列表):
    for 工作簿 in 工作簿列表 :
        try:
            工作簿.工作簿.remove(工作簿.工作簿['Sheet'])
        except :
            pass
        工作簿.保存文件()

def 写入表头(工作簿列表,表格名称,表头):
    for 工作簿 in 工作簿列表 :
        工作簿.工作簿[表格名称].append(表头)
    print('  写入表头完成!')


def 读取数据(文件路径):
    print('开始读取数据...')
    源文件 = openpyxl.load_workbook(文件路径, True)
    print('将处理工作簿最后3个表格')
    global 表格名称列表
    表格名称列表 = 源文件.sheetnames[-3:]
    print('要处理的3个表格为', 表格名称列表)
    表格列表 = [源文件[表格名称列表[0]], 源文件[表格名称列表[1]], 源文件[表格名称列表[2]]]
    print('读取数据完成。')
    return 表格列表


def 行数据转列表(rowData):
    rv = []
    for i in rowData:
        rv.append(i.value)
    return rv




def main():

    print('开始保存数据...')
    保存工作簿文件(工作簿列表)
    input("程序结束，输入任意值结束")

# main()

