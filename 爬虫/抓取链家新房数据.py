
from bs4 import BeautifulSoup
# Beautiful Soup 也是一个HTML/XML的解析器，主要的功能也是如何解析和提取 HTML/XML 数据。
# Beautiful Soup自动将输入文档转换为Unicode编码，输出文档转换为utf-8编码。
import requests
import time
import openpyxl
# import pandas as pd
headers = {
    'User-Agent': 'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.1; WOW64; Trident/6.0; SLCC2;.NET CLR 2.0.50727; .NET CLR 3.5.30729; .NET CLR 3.0.30729; InfoPath.3; .NET4.0C; .NET4.0E)',
    'Accept': 'image/webp,image/*,*/*;q=0.8',
    'Accept-Encoding': 'gzip, deflate',
    'Referer': 'http://www.baidu.com/link?url=_andhfsjjjKRgEWkj7i9cFmYYGsisrnm2A-TN3XZDQXxvGsM9k9ZZSnikW2Yds4s&amp;amp;wd=&amp;amp;eqid=c3435a7d00006bd600000003582bfd1f',
    'Connection': 'keep-alive'}

cityurl = 'https://tj.lianjia.com/ershoufang/'
r = requests.get(url=cityurl, headers=headers)

hlist = []
hlist.append(
    {'title': "楼盘名称", 'wuye': "物业类型", 'xiaoshouzhuangtai': "销售状态", 'location': "位置",
     'jishi': "房型", 'area': "面积", 'tag': "标签", 'price': "单价",
     'totalprice': "总价"})


def get_data(city='tj', house_type='loupan'):
    ''' 输入城市,返回要访问的网页\n
    天津:tj     \n
    新房:loupan     二手房: ershoufang   \n
    返回网页数据'''
    htmlinfo = bytes()  # 查询到的数据
    # 查询第1页
    url = 'https://' + city + '.lianjia.com/'+house_type+'/'
    r = requests.get(url=url, headers=headers)
    htmlinfo = r.content
    # 查询数据
    for i in range(2, 获取网页数量('')+1):  # 获取2-100页的数据
        a = (url + 'pg' + str(i) + '/')
        r = requests.get(url=a, headers=headers)
        htmlinfo = htmlinfo + r.content
        print('已获取到:',i,'/',获取网页数量(''),' ', a)
    print('已完成 获取网页数据')
    return htmlinfo

def w(mystr, filename='test.txt'):
    f = open(filename, 'w', encoding='utf-8')
    # f.write('mystr')
    f.write(str(mystr))
    f.close()

def 获取网页数量(cityurl):
    return 200
def listinfo(listhtml):
    '''处理获取的网页数据'''
    areasoup = BeautifulSoup(listhtml, 'html.parser')
    # 转换为格式化数据
    ljhouse = areasoup.find_all(
        'div', attrs={'class': 'resblock-desc-wrapper'})
    for house in ljhouse:
        loupantitle = house.find("div", attrs={"class": "resblock-name"})
        loupanname = loupantitle.a.get_text()
        loupantag = loupantitle.find_all("span")
        wuye = loupantag[0].get_text()
        xiaoshouzhuangtai = loupantag[1].get_text()
        location = house.find(
            "div", attrs={"class": "resblock-location"}).get_text()
        jishi = house.find("a", attrs={"class": "resblock-room"}).get_text()
        area = house.find("div", attrs={"class": "resblock-area"}).get_text()
        tag = house.find("div", attrs={"class": "resblock-tag"}).get_text()
        jiage = house.find("div", attrs={"class": "resblock-price"})
        price = jiage.find("div", attrs={"class": "main-price"}).get_text()
        total = jiage.find("div", attrs={"class": "second"})
        totalprice = "暂无"
        if total is not None:
            totalprice = total.get_text()
        h = {'title': loupanname, 'wuye': wuye, 'xiaoshouzhuangtai': xiaoshouzhuangtai, 'location': location.replace("\n", ""),
             'jishi': jishi.replace("\n", ""), 'area': area, 'tag': tag, 'price': price,
             'totalprice': totalprice}
        hlist.append(h)
        显示进度(house , ljhouse)

class 工作簿类():
    def __init__(self):
        self.工作簿 = openpyxl.Workbook()
        try:
            self.工作簿.remove(工作簿.工作簿['Sheet'])
        except:
            pass
        self.工作簿.create_sheet('链家新房')
        self.工作簿.close()

    def 写入一行数据(self, 行数据: list):
        self.工作簿['链家新房'].append(行数据)
    def 获取行数(self):
        return self.工作簿['链家新房'].max_row + 1

def 显示进度(i,l,d=10):
    id=l.index(i)
    n=len(l)
    if id % int(n /10) == 1 :
            print('已完成','{:.2f}'.format(100* id/ n ),'%')

if __name__ == '__main__':
    print('=='*15)
    print('获取链家房产数据', time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))
    print('=='*15)
    #input('请输入 要查询的城市：')
    city = 'tj'  # 目前固定为天津
    print('要查询的城市为:', '天津')
    # 获取网页原始数据
    print('开始查询数据')
    html_data = get_data()
    # 处理数据
    print('开始处理数据')
    listinfo(html_data)
    # 输出数据
    print('处理数据完成,开始输出数据')
    工作簿 = 工作簿类()
    for i in hlist:
        temp = list(i.values())
        工作簿.写入一行数据( temp)
        percent = '{:.2f}'.format(100*工作簿.获取行数()/len(hlist))
        显示进度(i,hlist)
    filename = '链家' + \
        time.strftime("%Y_%m_%d-%H_%M_%S", time.localtime())+'.xlsx'
    工作簿.工作簿.save(filename)
    print('end', '=='*15)
    input('按回车结束...')
